import csv
import mimetypes
import re
import time
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse

import httpx
from bs4 import BeautifulSoup
from pypdf import PdfReader

from ingestion.cleaning import clean_text


SPREADSHEET = Path("data/manual/Motif - Film Corpus.xlsx")
SPREADSHEET_CSV = Path("data/manual/manual_corpus_rows.csv")
MANUAL_DIR = Path("data/manual")
OUTPUT_DIR = Path("data/manual_extracted")
SOURCES_CSV = Path("data/manual_sources.csv")
REPORT_CSV = Path("data/manual_ingestion_report.csv")
MIN_WORDS = 80
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0.0.0 Safari/537.36"
)


@dataclass
class ManualRow:
    film_slug: str
    film_title: str
    year: int
    director: str
    slot_number: int
    target_source_type: str
    target_description: str
    status: str
    source_key: str
    source_title: str
    actual_source_type: str
    url: str


LOCAL_REPO_FILES = {
    ("fight-club", 7): MANUAL_DIR / "Transcript_How I Wrote Fight Club.txt",
    ("one-flew-over-the-cuckoos-nest", 8): MANUAL_DIR / "BJenningsBiopowerandtheLiberationistRomance2010.pdf",
    ("memento", 7): MANUAL_DIR / "Transcript_Memento - Interview with Christopher Nolan (2004).txt",
    ("prisoners", 2): MANUAL_DIR / "Transcript_PRISONERS Interviews Hugh Jackman, Jake Gyllenhaal, Paul Dano, Melissa Leo, Terrence Howard.txt",
    ("shawshank-redemption", 5): MANUAL_DIR / "Shawshank_Academic_Article.pdf",
    ("shawshank-redemption", 6): MANUAL_DIR / "Shawshank_Academic_2.pdf",
    ("shawshank-redemption", 7): MANUAL_DIR / "Shawshank_Video_Essay.pdf",
    ("fight-club", 5): MANUAL_DIR / "Fight_Club_Academic.pdf",
    ("one-flew-over-the-cuckoos-nest", 6): MANUAL_DIR / "One_Flew_Academic.pdf",
    ("se7en", 8): MANUAL_DIR / "Se7en_Review.pdf",
    ("silence-of-the-lambs", 1): MANUAL_DIR / "Silence_of_the_Lambs_Screenplay.pdf",
    ("silence-of-the-lambs", 3): MANUAL_DIR / "Silence_of_Lambs_Review.pdf",
    ("silence-of-the-lambs", 5): MANUAL_DIR / "Silence_of_Lambs_Academic.pdf",
    ("silence-of-the-lambs", 6): MANUAL_DIR / "Silence_of_Lambs_Essay.pdf",
    ("the-prestige", 3): MANUAL_DIR / "The_Prestige_IMDB.pdf",
    ("the-prestige", 5): MANUAL_DIR / "The_Prestige-Academic.pdf",
    ("the-prestige", 6): MANUAL_DIR / "Prestige_Essay.pdf",
    ("memento", 5): MANUAL_DIR / "Memento_Academic.pdf",
    ("taxi-driver", 2): MANUAL_DIR / "Taxi_Driver_People_Magazine.pdf",
    ("taxi-driver", 5): MANUAL_DIR / "Taxi_Driver_review.pdf",
    ("shutter-island", 3): MANUAL_DIR / "Shutter_Island_Review.pdf",
    ("black-swan", 6): MANUAL_DIR / "Black_Swan_Academic.pdf",
    ("black-swan", 7): MANUAL_DIR / "Black_Swan_Review.pdf",
    ("sixth-sense", 6): MANUAL_DIR / "Sixth_Sense_Academic.pdf",
    ("sixth-sense", 7): MANUAL_DIR / "Sixth_Sense_review.pdf",
    ("prisoners", 3): MANUAL_DIR / "Prisoners_Blog.pdf",
    ("prisoners", 5): MANUAL_DIR / "Prisoners_Craft_Article.pdf",
    ("prisoners", 7): MANUAL_DIR / "Prisoners_review.pdf",
    ("gone-girl", 5): MANUAL_DIR / "Gone_Girl_Academic.pdf",
    ("gone-girl", 6): MANUAL_DIR / "Gone_Girl_Essay.pdf",
    ("donnie-darko", 5): MANUAL_DIR / "Donnie_Darko_Academic.pdf",
    ("donnie-darko", 6): MANUAL_DIR / "Donnie_Darko_Academic2.pdf",
    ("donnie-darko", 7): MANUAL_DIR / "Donnie_Darko_Review.pdf",
    ("the-machinist", 5): MANUAL_DIR / "The_Machinist_Academic.pdf",
    ("the-machinist", 6): MANUAL_DIR / "Machinist_Academic2.pdf",
    ("mulholland-drive", 2): MANUAL_DIR / "Mulholland_Dr_Interview.pdf",
    ("mulholland-drive", 5): MANUAL_DIR / "Mulholland_Drive_Academic.pdf",
    ("mulholland-drive", 7): MANUAL_DIR / "Mulholland_Drive_review.pdf",
    ("truman-show", 5): MANUAL_DIR / "Truman_Show_Academic.pdf",
    ("truman-show", 6): MANUAL_DIR / "Truman_Show_Essay.pdf",
}


LOCAL_SOURCE_TYPE_OVERRIDES = {
    ("one-flew-over-the-cuckoos-nest", 6): "academic",
    ("se7en", 8): "review",
    ("silence-of-the-lambs", 3): "review",
    ("the-prestige", 3): "festival_qa",
    ("taxi-driver", 2): "interview",
    ("taxi-driver", 5): "review",
    ("shutter-island", 3): "review",
    ("black-swan", 7): "review",
    ("sixth-sense", 7): "review",
    ("prisoners", 3): "festival_qa",
    ("prisoners", 7): "review",
    ("donnie-darko", 7): "review",
    ("mulholland-drive", 2): "interview",
    ("mulholland-drive", 7): "review",
}


def slugify(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def domain_for(url: str) -> str:
    return urlparse(url).netloc.lower().removeprefix("www.")


def read_pdf(path: Path) -> str:
    reader = PdfReader(str(path), strict=False)
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def html_to_text(content: str, url: str) -> tuple[str, str]:
    soup = BeautifulSoup(content, "html.parser")
    for tag in soup(["script", "style", "nav", "footer", "header", "aside", "form", "noscript"]):
        tag.decompose()
    title = soup.find("title")
    title_text = title.get_text(" ", strip=True) if title else url
    article = soup.find("article") or soup.find("main") or soup
    text = article.get_text("\n", strip=True)
    return title_text, text


def fetch_url(url: str) -> tuple[str, str]:
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,application/pdf,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    with httpx.Client(headers=headers, timeout=30, follow_redirects=True, verify=False) as client:
        response = client.get(url)
        response.raise_for_status()
        content_type = response.headers.get("content-type", "").lower()
        suffix = Path(urlparse(str(response.url)).path).suffix.lower()
        if "pdf" in content_type or suffix == ".pdf":
            temp_path = OUTPUT_DIR / "_tmp_download.pdf"
            temp_path.write_bytes(response.content)
            return Path(urlparse(url).path).name or url, read_pdf(temp_path)
        if "text/plain" in content_type or suffix == ".txt":
            return Path(urlparse(url).path).name or url, response.text
        return html_to_text(response.text, url)


def read_local(path: Path) -> tuple[str, str]:
    if path.suffix.lower() == ".pdf":
        return path.stem, read_pdf(path)
    return path.stem, path.read_text(encoding="utf-8", errors="ignore")


def infer_source_type(target: str, url: str, slot_number: int) -> str:
    lowered = url.lower()
    domain = domain_for(url)
    if url.strip().lower() == "in repo":
        if slot_number == 2:
            return "interview"
        if slot_number == 7:
            return "video_essay_transcript"
        return target or "other"
    if any(token in lowered for token in ("script", "screenplay", "dailyscript", "imsdb", "scriptslug")):
        return "screenplay"
    if "catalog.afi.com" in domain:
        return "production_notes"
    if domain in {"rogerebert.com", "variety.com"}:
        return "review"
    if domain == "theguardian.com" and "review" in lowered:
        return "review"
    if domain in {"theasc.com", "studiobinder.com", "goldderby.com"}:
        return "craft_article"
    if any(token in domain for token in ("digitalcommons", "scholarworks", "purdue.edu", "semanticscholar", "jstor", "mdpi", "pmc.ncbi", "nottingham.ac.uk", "euppublishing", "academypublication")):
        return "academic"
    if domain in {"bfi.org.uk", "criterion.com", "sensesofcinema.com", "offscreen.com", "bfidatadigipres.github.io"}:
        return "educational_essay"
    if any(token in lowered for token in ("interview", "q-a", "qa", "conversation")):
        return "interview" if target == "interview" else target
    return target or "other"


def primary_for(source_type: str) -> bool:
    return source_type in {
        "screenplay",
        "interview",
        "festival_qa",
        "production_notes",
        "director_commentary",
        "cast_interview",
    }


def credibility_for(source_type: str, url: str) -> float:
    domain = domain_for(url)
    if source_type in {"screenplay", "interview", "production_notes", "festival_qa"}:
        return 0.92
    if source_type == "academic":
        return 0.9
    if source_type in {"educational_essay", "craft_article"}:
        return 0.86
    if source_type == "review":
        return 0.82
    if domain:
        return 0.78
    return 0.7


def load_rows() -> list[ManualRow]:
    if SPREADSHEET_CSV.exists():
        with SPREADSHEET_CSV.open(newline="", encoding="utf-8") as handle:
            raw_rows = list(csv.reader(handle))[1:]
    else:
        from openpyxl import load_workbook

        workbook = load_workbook(SPREADSHEET, data_only=True)
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    rows = []
    for row in raw_rows:
        if not row[0] or not row[11]:
            continue
        film_slug = str(row[0]).strip()
        slot_number = int(float(row[4]))
        target_type = str(row[5] or "").strip()
        url = str(row[11]).strip()
        actual_type = str(row[10] or "").strip() or infer_source_type(target_type, url, slot_number)
        actual_type = LOCAL_SOURCE_TYPE_OVERRIDES.get((film_slug, slot_number), actual_type)
        source_key = str(row[8] or "").strip()
        if not source_key:
            source_key = f"{film_slug}-{slot_number:02d}-{actual_type}-{slugify(domain_for(url) or 'local')}"
        rows.append(
            ManualRow(
                film_slug=film_slug,
                film_title=str(row[1]).strip(),
                year=int(float(row[2])),
                director=str(row[3]).strip(),
                slot_number=slot_number,
                target_source_type=target_type,
                target_description=str(row[6] or "").strip(),
                status=str(row[7] or "").strip(),
                source_key=source_key,
                source_title=str(row[9] or "").strip(),
                actual_source_type=actual_type,
                url=url,
            )
        )
    return rows


def build() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    source_rows = []
    report_rows = []
    for row in load_rows():
        local_repo_path = LOCAL_REPO_FILES.get((row.film_slug, row.slot_number))
        try:
            if local_repo_path:
                title, raw_text = read_local(local_repo_path)
                original_url = "In repo"
            else:
                time.sleep(0.15)
                title, raw_text = fetch_url(row.url)
                original_url = row.url
            cleaned = clean_text(raw_text)
            word_count = len(cleaned.split())
            if word_count < MIN_WORDS:
                raise ValueError(f"extracted text too short ({word_count} words)")
            source_title = row.source_title or title or f"{row.film_title} source {row.slot_number}"
            output_path = OUTPUT_DIR / f"{row.source_key}.txt"
            output_path.write_text(
                f"{source_title}\n{original_url}\n\n{cleaned}\n",
                encoding="utf-8",
            )
            source_rows.append(
                {
                    "film_slug": row.film_slug,
                    "source_key": row.source_key,
                    "title": source_title,
                    "author": "",
                    "publisher": domain_for(row.url) or "local upload",
                    "source_type": row.actual_source_type,
                    "url": "" if original_url == "In repo" else row.url,
                    "publication_date": "",
                    "is_primary": str(primary_for(row.actual_source_type)).lower(),
                    "credibility_score": f"{credibility_for(row.actual_source_type, row.url):.2f}",
                    "local_path": str(output_path),
                    "notes": f"Manual corpus row {row.slot_number}. Spreadsheet target type: {row.target_source_type}. {row.target_description}",
                }
            )
            report_rows.append(
                {
                    "film_slug": row.film_slug,
                    "slot_number": row.slot_number,
                    "source_key": row.source_key,
                    "status": "ok",
                    "source_type": row.actual_source_type,
                    "word_count": word_count,
                    "url": row.url,
                    "message": "",
                }
            )
        except Exception as exc:
            report_rows.append(
                {
                    "film_slug": row.film_slug,
                    "slot_number": row.slot_number,
                    "source_key": row.source_key,
                    "status": "failed",
                    "source_type": row.actual_source_type,
                    "word_count": 0,
                    "url": row.url,
                    "message": str(exc),
                }
            )
    fieldnames = [
        "film_slug",
        "source_key",
        "title",
        "author",
        "publisher",
        "source_type",
        "url",
        "publication_date",
        "is_primary",
        "credibility_score",
        "local_path",
        "notes",
    ]
    with SOURCES_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(source_rows)
    with REPORT_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(report_rows[0].keys()))
        writer.writeheader()
        writer.writerows(report_rows)
    ok_count = sum(1 for item in report_rows if item["status"] == "ok")
    print(f"wrote {SOURCES_CSV} with {len(source_rows)} rows")
    print(f"wrote {REPORT_CSV}: ok={ok_count} failed={len(report_rows) - ok_count}")


if __name__ == "__main__":
    build()
